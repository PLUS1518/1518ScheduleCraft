from telegram import (
    Update,
    ReplyKeyboardRemove  # <-- Добавьте эту строку
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    ConversationHandler,
    filters
)
import logging
logger = logging.getLogger(__name__)
import random
import sys
import psutil
import os
import asyncio
import json
import aiosqlite
import openpyxl
from openpyxl.styles import Font
from datetime import datetime, timedelta
from collections import defaultdict
from contextlib import contextmanager

# Конфигурация
TOKEN = '7762421006:AAEVrYZMgI3XW__EPWWS3fehEEkrJjDFeOo'
ADMIN_CODE = "1518"
DATABASE_NAME = 'schedules.db'
ADMIN_CHAT_ID = 805086834

# Состояния диалога
(
    SUBJECTS,
    INPUT_SUBJECT,
    TEACHERS,
    TEACHER_SUBJECTS,
    TEACHER_TIME,
    CLASSES,
    CLASS_MAX_LESSONS,
    CLASS_MAX_LESSONS_VALUE,
    CLASS_GROUPS_INPUT,
    SCHEDULE,
) = range(10)

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Глобальные переменные
ADMIN_COMMANDS = ['/test', '/crash', '/restart']
ADMIN_CHAT_ID = 805086834
RESTART_LIMIT = 10
TIME_WINDOW = timedelta(minutes=10)
user_restarts = defaultdict(list)

async def log_state_transition(from_state, to_state, user_id):
    logger.info(f"User {user_id}: Transition {from_state} → {to_state}")

async def init_db():
    """Инициализация структуры базы данных"""
    async with aiosqlite.connect(DATABASE_NAME) as conn:
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                username TEXT,
                created TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        await conn.execute('''
            CREATE TABLE IF NOT EXISTS subjects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                name TEXT,
                FOREIGN KEY(user_id) REFERENCES users(user_id)
            )
        ''')

        await conn.execute('''
            CREATE TABLE IF NOT EXISTS teachers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                fio TEXT,
                subjects TEXT,
                preferred_times TEXT,
                schedule TEXT,
                lessons_count INTEGER DEFAULT 0,
                FOREIGN KEY(user_id) REFERENCES users(user_id)
            )
        ''')

        await conn.execute('''
            CREATE TABLE IF NOT EXISTS classes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                name TEXT,
                max_lessons INTEGER,
                lessons TEXT,
                FOREIGN KEY(user_id) REFERENCES users(user_id)
            )
        ''')

        # Индексы для ускорения запросов
        await conn.execute('CREATE INDEX IF NOT EXISTS idx_subjects_user ON subjects(user_id)')
        await conn.execute('CREATE INDEX IF NOT EXISTS idx_teachers_user ON teachers(user_id)')
        await conn.execute('CREATE INDEX IF NOT EXISTS idx_classes_user ON classes(user_id)')
        await conn.commit()


async def admin_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Запрос кода администратора"""
    await update.message.reply_text("Введите код администратора:")


async def handle_admin_code(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Безопасная обработка кода администратора"""
    try:
        user_id = update.effective_user.id
        input_code = update.message.text.strip()

        if input_code == ADMIN_CODE:
            # Инициализация флага администратора
            context.user_data['is_admin'] = True
            await update.message.reply_text("✅ Вы авторизованы как администратор!")
        else:
            await update.message.reply_text("❌ Неверный код доступа")

        # Явный сброс состояния
        if 'conversation' in context.chat_data:
            del context.chat_data['conversation']

    except Exception as e:
        logger.error(f"Ошибка авторизации: {e}")
        await update.message.reply_text("⚠️ Произошла ошибка при обработке кода")

async def check_admin(context: ContextTypes.DEFAULT_TYPE, user_id: int) -> bool:
    """Проверка прав администратора"""
    # Проверяем ID чата и флаг в user_data
    return user_id == ADMIN_CHAT_ID or context.user_data.get('is_admin', False)


# Административные команды
async def test_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Тестовая генерация расписания"""
    user_id = update.effective_user.id

    # Очистка предыдущих данных
    async with aiosqlite.connect(DATABASE_NAME) as conn:
        await conn.execute('DELETE FROM subjects WHERE user_id = ?', (user_id,))
        await conn.execute('DELETE FROM teachers WHERE user_id = ?', (user_id,))
        await conn.execute('DELETE FROM classes WHERE user_id = ?', (user_id,))
        await conn.commit()

    # Создание тестовых данных
    subjects = [
        "Русский язык", "Литература", "Алгебра", "Геометрия",
        "Физика", "Химия", "Биология", "История",
        "Обществознание", "География", "Информатика",
        "Английский язык", "Физкультура", "ОБЖ",
        "Технология", "Музыка", "ИЗО"
    ]

    # Генерация случайных преподавателей
    surnames = ["Иванов", "Петров", "Сидоров", "Кузнецова", "Смирнова", "Васильев",
                "Попова", "Соколова", "Михайлов", "Новикова", "Федорова", "Морозов",
                "Волков", "Алексеева", "Лебедева", "Семенова", "Егорова", "Павлова",
                "Козлова", "Степанова"]
    initials = ["А.А.", "И.И.", "С.С.", "М.М.", "Д.Д.", "О.О.", "Н.Н.", "В.В.", "П.П.", "К.К."]

    teachers = []
    used_names = set()
    # Генерируем 30 учителей
    for _ in range(30):
        while True:
            surname = random.choice(surnames)
            initial = random.choice(initials)
            fio = f"{surname} {initial}"
            if fio not in used_names:
                used_names.add(fio)
                break
        # Выбираем 1-2 предмета, исключая дубликаты
        teacher_subjects = random.sample(subjects, k=random.randint(1, 2))
        teacher_subjects = list(set(teacher_subjects))  # Удаляем дубликаты, если есть
        # Предпочтительные дни: 0-4 (пн-пт)
        preferred_times = random.sample(range(5), k=random.randint(0, 3))
        teachers.append({
            'fio': fio,
            'subjects': teacher_subjects,
            'preferred_times': preferred_times,
            'schedule': [[False] * 8 for _ in range(5)]  # 5 дней, 8 уроков
        })

    # Проверка покрытия предметов и добавление недостающих
    covered_subjects = set()
    for teacher in teachers:
        covered_subjects.update(teacher['subjects'])
    missing_subjects = set(subjects) - covered_subjects
    for sub in missing_subjects:
        # Добавляем учителя для каждого недостающего предмета
        while True:
            surname = random.choice(surnames)
            initial = random.choice(initials)
            fio = f"{surname} {initial}"
            if fio not in used_names:
                used_names.add(fio)
                break
        teachers.append({
            'fio': fio,
            'subjects': [sub],
            'preferred_times': [],
            'schedule': [[False] * 8 for _ in range(5)]
        })

    # Создание 11 классов
    class_names = ["5А", "5Б", "6А", "6Б", "7А", "7Б", "8А", "9А", "9Б", "10А", "11А"]
    classes_data = []
    for class_name in class_names:
        max_lessons = random.randint(5, 8)
        # Выбираем случайные предметы для класса
        num_subjects = random.randint(10, len(subjects))
        class_subjects = random.sample(subjects, k=num_subjects)
        lessons = {}
        total = 0
        max_total = max_lessons * 5  # Максимальное количество уроков в неделю
        for sub in class_subjects:
            if total >= max_total:
                break
            max_possible = min(6, max_total - total)
            count = random.randint(1, max_possible)
            lessons[sub] = count
            total += count
        classes_data.append({
            'name': class_name,
            'max_lessons': max_lessons,
            'lessons': lessons
        })

    # Сохранение тестовых данных в БД
    async with aiosqlite.connect(DATABASE_NAME) as conn:
        # Сохраняем предметы
        for subject in subjects:
            await conn.execute('INSERT INTO subjects (user_id, name) VALUES (?, ?)', (user_id, subject))

        # Сохраняем учителей
        for teacher in teachers:
            await conn.execute('''
                INSERT INTO teachers (user_id, fio, subjects, preferred_times, schedule)
                VALUES (?, ?, ?, ?, ?)
            ''', (
                user_id,
                teacher['fio'],
                json.dumps(teacher['subjects']),
                json.dumps(teacher['preferred_times']),
                json.dumps(teacher['schedule'])
            ))

        # Сохраняем классы
        for class_data in classes_data:
            await conn.execute('''
                INSERT INTO classes (user_id, name, max_lessons, lessons)
                VALUES (?, ?, ?, ?)
            ''', (
                user_id,
                class_data['name'],
                class_data['max_lessons'],
                json.dumps(class_data['lessons'])
            ))
        await conn.commit()

    # Генерация расписания
    data = await get_user_data(user_id)
    schedule = await create_schedule(data['classes'])
    schedule, all_placed = await fill_schedule(data['teachers'], data['classes'], schedule)

    # Создание Excel-файла
    filename = f"test_schedule_{user_id}.xlsx"
    await create_excel_schedule(schedule, data['classes'], filename)

    # Отправка файла
    with open(filename, "rb") as f:
        await update.message.reply_document(
            document=f,
            filename=filename,
            caption="✅ Тестовое расписание сгенерировано"
        )

    os.remove(filename)


async def restart_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Перезагрузка бота"""
    user_id = update.effective_user.id
    if not await check_admin(context, user_id):
        await update.message.reply_text("🚫 Требуются права администратора!")
        return

    await update.message.reply_text("🔄 Перезагрузка...")
    os.execl(sys.executable, sys.executable, *sys.argv)


async def crash_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Искусственная ошибка"""
    user_id = update.effective_user.id
    if not await check_admin(context, user_id):
        await update.message.reply_text("🚫 Требуются права администратора!")
        return

    raise RuntimeError("Искусственный краш по запросу администратора")


async def get_user_data(user_id: int) -> dict:
    """Получение всех данных пользователя из БД"""
    data = {'subjects': [], 'teachers': {}, 'classes': {}}

    async with aiosqlite.connect(DATABASE_NAME) as conn:
        # Получаем предметы
        cursor = await conn.execute('SELECT name FROM subjects WHERE user_id = ?', (user_id,))
        data['subjects'] = [row[0] async for row in cursor]

        # Получаем учителей
        cursor = await conn.execute('''
            SELECT id, fio, subjects, preferred_times, schedule, lessons_count 
            FROM teachers WHERE user_id = ?
        ''', (user_id,))

        async for row in cursor:
            data['teachers'][row[0]] = {
                'fio': row[1],
                'subjects': json.loads(row[2]),
                'preferred_times': json.loads(row[3]),
                'schedule': json.loads(row[4]),
                'lessons_count': row[5]
            }

        # Получаем классы
        cursor = await conn.execute('''
            SELECT name, max_lessons, lessons FROM classes WHERE user_id = ?
        ''', (user_id,))

        async for row in cursor:
            data['classes'][row[0]] = {
                'max_lessons_per_day': row[1],
                'lessons': json.loads(row[2])
            }

    return data


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка команды /start"""
    user = update.effective_user

    async with aiosqlite.connect(DATABASE_NAME) as conn:
        await conn.execute('INSERT OR IGNORE INTO users (user_id, username) VALUES (?, ?)',
                           (user.id, user.username))
        await conn.commit()

    await update.message.reply_text(
        "Привет! Я бот для составления расписания.\n"
        "Начните с команды /setup, чтобы задать данные.\n"
        "Или /help для помощи."
    )
    return ConversationHandler.END


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Показ справки"""
    help_text = [
        "Основные команды:",
        "/start - начать работу ▶️",
        "/setup - ввод данных 📋",
        "/schedule - создать расписание🗓️",
        "/new - очистить все данные 🆕",
        "/help - эта справка 🆘",
        "/test - тестовый режим📍",
        "/admin - войти в систему администрирования 🔐",
        "Поддержка - @Pluss_pr 🗨️",
    ]

    if context.user_data.get('is_admin'):
        help_text.extend([
            "\nАдминские команды 🕵️‍♂️:",
            "/crash - экстренное завершение 🛑",
            "/restart - перезапуск бота 🔁",
            "/log - получение логов (недоступно ⛔)",
            "/stats - получение статистики (недоступно ⛔)",
            "/ads - управление доступом (недоступно ⛔)"
        ])

    await update.message.reply_text("\n".join(help_text))


async def setup(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if context.user_data:
        await new_command(update, context)  # Форсированный сброс

    # Проверка пустой БД
    async with aiosqlite.connect(DATABASE_NAME) as conn:
        cursor = await conn.execute('SELECT COUNT(*) FROM subjects WHERE user_id = ?',
                                    (update.effective_user.id,))
        count = (await cursor.fetchone())[0]
        if count > 0:
            await new_command(update, context)
    user_id = update.effective_user.id
    if context.user_data.get('__conversation__'):
        await context.application.dispatcher.update_queue.put(
            (update.update_id, lambda: None)
        )
        context.user_data['__conversation__'] = None
    # Очистка предыдущих данных и состояния
    context.user_data.clear()  # Сбрасываем все переменные состояния

    async with aiosqlite.connect(DATABASE_NAME) as conn:
        await conn.execute('DELETE FROM subjects WHERE user_id = ?', (user_id,))
        await conn.execute('DELETE FROM teachers WHERE user_id = ?', (user_id,))
        await conn.execute('DELETE FROM classes WHERE user_id = ?', (user_id,))
        await conn.commit()

    await update.message.reply_text("Введите количество предметов в школе:")
    return SUBJECTS


async def input_subjects(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        num_subjects = int(update.message.text)
        if num_subjects < 1:
            raise ValueError
        context.user_data['subject_counter'] = 0
        context.user_data['num_subjects'] = num_subjects
        await update.message.reply_text(f"Введите название предмета 1:")
        return INPUT_SUBJECT
    except Exception as e:
        logger.error(f"Ошибка в input_subjects: {e}")
        await update.message.reply_text("❌ Ошибка ввода. Введите положительное число.")
        return SUBJECTS


async def create_excel_schedule(schedule, classes, filename):
    """Создание Excel-файла с расписанием"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Удаляем дефолтный лист

    for class_name in schedule:
        # Создаем новый лист для каждого класса
        ws = wb.create_sheet(title=class_name[:31])  # Ограничение длины названия листа

        # Заголовки
        headers = ["Урок"] + ["Пн", "Вт", "Ср", "Чт", "Пт"]
        ws.append(headers)

        # Форматирование заголовков
        for col in range(1, 6):
            cell = ws.cell(row=1, column=col + 1)
            cell.font = Font(bold=True)

        # Заполняем данные
        max_lessons = classes[class_name]['max_lessons_per_day']
        for lesson_num in range(max_lessons):
            row = [lesson_num + 1]
            for day in range(5):
                cell_value = schedule[class_name][lesson_num][day] or "Свободно"
                row.append(cell_value)
            ws.append(row)

        # Настройка ширины столбцов
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    wb.save(filename)



async def input_subjects_next(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработчик ввода предметов с проверкой счетчика"""
    # Проверка и инициализация счетчика
    if 'subject_counter' not in context.user_data:
        context.user_data['subject_counter'] = 0

    user_id = update.effective_user.id
    subject = update.message.text.strip()

    async with aiosqlite.connect(DATABASE_NAME) as conn:
        await conn.execute('INSERT INTO subjects (user_id, name) VALUES (?, ?)',
                           (user_id, subject))
        await conn.commit()

    context.user_data['subject_counter'] += 1

    if context.user_data['subject_counter'] < context.user_data['num_subjects']:
        await update.message.reply_text(
            f"Введите название предмета {context.user_data['subject_counter'] + 1}:"
        )
        return INPUT_SUBJECT

    await update.message.reply_text("Введите количество учителей:")
    return TEACHERS


async def input_subjects_next(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Сохранение предмета в БД"""
    user_id = update.effective_user.id
    subject = update.message.text.strip()

    async with aiosqlite.connect(DATABASE_NAME) as conn:
        await conn.execute('INSERT INTO subjects (user_id, name) VALUES (?, ?)',
                           (user_id, subject))
        await conn.commit()

    context.user_data['subject_counter'] += 1

    if context.user_data['subject_counter'] < context.user_data['num_subjects']:
        await update.message.reply_text(
            f"Введите название предмета {context.user_data['subject_counter'] + 1}:"
        )
        return INPUT_SUBJECT

    await update.message.reply_text("Введите количество учителей:")
    return TEACHERS


async def input_teachers(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка количества учителей"""
    try:
        num = int(update.message.text)
        if num < 1:
            raise ValueError
        context.user_data['num_teachers'] = num
        context.user_data['teacher_counter'] = 0
        await update.message.reply_text("Введите ФИО учителя 1:")
        return TEACHER_SUBJECTS
    except ValueError:
        await update.message.reply_text("Введите положительное число!")
        return TEACHERS


async def input_teachers_subjects(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка данных учителя"""
    user_id = update.effective_user.id
    text = update.message.text.strip()

    if 'current_teacher' not in context.user_data:
        context.user_data['current_teacher'] = {
            'fio': text,
            'subjects': [],
            'preferred_times': []
        }
        await update.message.reply_text("Введите предметы для этого учителя (через запятую):")
        return TEACHER_SUBJECTS

    # Получаем список предметов из БД
    async with aiosqlite.connect(DATABASE_NAME) as conn:
        cursor = await conn.execute('SELECT name FROM subjects WHERE user_id = ?', (user_id,))
        existing_subjects = {row[0] async for row in cursor}

    # Проверяем введенные предметы
    input_subjects = [s.strip() for s in text.split(',')]
    invalid = [s for s in input_subjects if s not in existing_subjects]

    if invalid:
        await update.message.reply_text(f"Ошибка: предметы {', '.join(invalid)} не найдены")
        return TEACHER_SUBJECTS

    context.user_data['current_teacher']['subjects'] = input_subjects
    await update.message.reply_text(
        "Введите предпочтительные дни (1-5 через пробел) или 'нет':"
    )
    return TEACHER_TIME


async def input_teachers_time(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Сохранение учителя в БД"""
    user_id = update.effective_user.id
    text = update.message.text.strip().lower()
    teacher = context.user_data['current_teacher']

    # Обработка предпочтительных дней
    preferred_times = []
    if text != 'нет':
        try:
            days = list(map(int, text.split()))
            if any(not (1 <= d <= 5) for d in days):
                raise ValueError
            preferred_times = [d - 1 for d in days]  # Конвертируем в 0-based
        except ValueError:
            await update.message.reply_text("Некорректный формат дней!")
            return TEACHER_TIME

    # Сохранение в БД
    async with aiosqlite.connect(DATABASE_NAME) as conn:
        await conn.execute('''
            INSERT INTO teachers 
            (user_id, fio, subjects, preferred_times, schedule)
            VALUES (?, ?, ?, ?, ?)
        ''', (
            user_id,
            teacher['fio'],
            json.dumps(teacher['subjects']),
            json.dumps(preferred_times),
            json.dumps([[False] * 8 for _ in range(5)])  # Пустое расписание
        ))
        await conn.commit()

    context.user_data['teacher_counter'] += 1
    del context.user_data['current_teacher']

    if context.user_data['teacher_counter'] < context.user_data['num_teachers']:
        await update.message.reply_text(
            f"Введите ФИО учителя {context.user_data['teacher_counter'] + 1}:"
        )
        return TEACHER_SUBJECTS

    await update.message.reply_text("Введите количество классов:")
    return CLASSES


async def input_classes(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка количества классов"""
    try:
        num = int(update.message.text)
        if num < 1:
            raise ValueError
        context.user_data['num_classes'] = num
        context.user_data['class_counter'] = 0
        await update.message.reply_text("Введите букву класса 1:")
        return CLASS_MAX_LESSONS
    except ValueError:
        await update.message.reply_text("Введите положительное число!")
        return CLASSES


async def input_classes_max_lessons(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработчик ввода названия класса и инициализации структуры данных"""
    try:
        # Получаем и нормализуем название класса
        class_name = update.message.text.strip().upper()

        if not class_name:
            await update.message.reply_text("❌ Название класса не может быть пустым!")
            return CLASS_MAX_LESSONS  # Повторно запрашиваем ввод

        # Инициализация структуры данных для нового класса
        context.user_data['current_class'] = {
            'name': class_name,
            'max_lessons': None,
            'lessons': {}
        }

        # Сбрасываем индекс предметов для нового класса
        context.user_data['subject_index'] = 0

        # Получаем список предметов для проверки
        user_id = update.effective_user.id
        async with aiosqlite.connect(DATABASE_NAME) as conn:
            cursor = await conn.execute('SELECT name FROM subjects WHERE user_id = ?', (user_id,))
            subjects_exist = any(await cursor.fetchall())

        if not subjects_exist:
            await update.message.reply_text("❌ Сначала добавьте предметы через /setup!")
            return ConversationHandler.END

        await update.message.reply_text(
            f"Введите максимальное количество уроков в день для класса {class_name} (1-8):"
        )
        return CLASS_MAX_LESSONS_VALUE

    except Exception as e:
        logger.error(f"Ошибка в input_classes_max_lessons: {str(e)}")
        await update.message.reply_text("⚠️ Произошла ошибка. Начните заново с /setup")
        return ConversationHandler.END


async def input_classes_max_lessons_value(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработчик ввода максимального количества уроков"""
    try:
        max_lessons = int(update.message.text)
        if not 1 <= max_lessons <= 8:
            raise ValueError

        # Проверяем наличие структуры класса
        if 'current_class' not in context.user_data:
            context.user_data['current_class'] = {}

        # Сохраняем значение
        context.user_data['current_class']['max_lessons'] = max_lessons

        # Дополнительная проверка
        logger.debug(f"Сохраненные данные класса: {context.user_data['current_class']}")

        # Запрос количества уроков
        await update.message.reply_text(
            f"Введите количество уроков для первого предмета:"
        )
        return SCHEDULE

    except ValueError:
        await update.message.reply_text("❌ Введите число от 1 до 8!")
        return CLASS_MAX_LESSONS_VALUE


async def input_classes_groups(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработчик ввода количества уроков для предмета"""
    try:
        user_id = update.effective_user.id
        text = update.message.text.strip().lower()

        # Проверка наличия необходимых данных в контексте
        if 'current_class' not in context.user_data:
            await update.message.reply_text("❌ Ошибка конфигурации! Начните заново с /setup")
            return ConversationHandler.END

        current_class = context.user_data['current_class']

        # Валидация структуры данных класса
        if 'name' not in current_class or not current_class.get('name'):
            await update.message.reply_text("❌ Не указано название класса!")
            return ConversationHandler.END

        if 'max_lessons' not in current_class or not current_class.get('max_lessons'):
            await update.message.reply_text("❌ Не указано максимальное количество уроков!")
            return ConversationHandler.END

        # Получение списка предметов из БД
        async with aiosqlite.connect(DATABASE_NAME) as conn:
            cursor = await conn.execute('SELECT name FROM subjects WHERE user_id = ?', (user_id,))
            subjects = [row[0] async for row in cursor]

        if not subjects:
            await update.message.reply_text("❌ Нет доступных предметов! Начните с /setup")
            return ConversationHandler.END

        # Получение текущего индекса предмета
        current_index = context.user_data.get('subject_index', 0)

        # Проверка корректности индекса
        if current_index >= len(subjects):
            await update.message.reply_text("⚠️ Ошибка: Несоответствие количества предметов")
            return ConversationHandler.END

        current_subject = subjects[current_index]

        # Обработка ввода пользователя
        if text == 'нет':
            lesson_count = 0
        else:
            try:
                lesson_count = int(text)
                if lesson_count < 0:
                    raise ValueError
            except ValueError:
                await update.message.reply_text("❌ Некорректное число! Введите положительное число или 'нет'")
                return SCHEDULE

        # Сохранение данных
        context.user_data['current_class']['lessons'][current_subject] = lesson_count
        context.user_data['subject_index'] = current_index + 1

        # Проверка оставшихся предметов
        if context.user_data['subject_index'] < len(subjects):
            next_subject = subjects[context.user_data['subject_index']]
            await update.message.reply_text(
                f"Введите количество уроков для предмета '{next_subject}' "
                "(или 'нет' если нет уроков):"
            )
            return SCHEDULE

        # Сохранение класса в БД
        try:
            async with aiosqlite.connect(DATABASE_NAME) as conn:
                await conn.execute('''
                    INSERT INTO classes 
                    (user_id, name, max_lessons, lessons)
                    VALUES (?, ?, ?, ?)
                ''', (
                    user_id,
                    current_class['name'],
                    current_class['max_lessons'],
                    json.dumps(current_class['lessons'])
                ))
                await conn.commit()
        except Exception as e:
            logger.error(f"Ошибка сохранения класса: {str(e)}")
            await update.message.reply_text("⚠️ Ошибка сохранения данных. Попробуйте снова")
            return SCHEDULE

        # Переход к следующему классу
        context.user_data['class_counter'] += 1
        if context.user_data['class_counter'] < context.user_data.get('num_classes', 0):
            await update.message.reply_text(
                f"Введите букву класса {context.user_data['class_counter'] + 1}:"
            )
            return CLASS_MAX_LESSONS

        await update.message.reply_text(
            "✅ Все классы добавлены! Используйте /schedule для генерации расписания"
        )
        return ConversationHandler.END

    except KeyError as e:
        logger.error(f"KeyError: Отсутствует ключ в контексте - {str(e)}")
        await update.message.reply_text("⚠️ Ошибка конфигурации. Начните заново с /setup")
        return ConversationHandler.END

    except ValueError as e:
        logger.error(f"ValueError: Некорректные данные - {str(e)}")
        await update.message.reply_text("❌ Ошибка ввода данных. Проверьте вводимые значения")
        return SCHEDULE

    except Exception as e:
        logger.error(f"Неожиданная ошибка: {str(e)}")
        await update.message.reply_text("⚠️ Критическая ошибка. Пожалуйста, начните заново с /setup")
        return ConversationHandler.END
    
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Отмена текущего диалога"""
    await update.message.reply_text("Диалог отменён. Начните заново с /setup.")
    context.user_data.clear()
    return ConversationHandler.END
    
    
async def schedule_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Генерация и отправка расписания"""
    user_id = update.effective_user.id
    data = await get_user_data(user_id)

    if not data.get('classes'):
        await update.message.reply_text("Сначала введите данные через /setup")
        return

    schedule = await create_schedule(data['classes'])
    schedule, all_placed = await fill_schedule(data['teachers'], data['classes'], schedule)

    if not all_placed:
        await update.message.reply_text("⚠️ Не удалось разместить все уроки!")

    filename = f"schedule_{user_id}.xlsx"
    await create_excel_schedule(schedule, data['classes'], filename)

    with open(filename, "rb") as f:
        await update.message.reply_document(
            document=f,
            filename=filename,
            caption="✅ Расписание готово!"
        )

    os.remove(filename)


async def create_schedule(classes):
    """Создание пустого расписания"""
    schedule = {}
    for class_name, data in classes.items():
        schedule[class_name] = [
            ["" for _ in range(5)]  # 5 дней
            for _ in range(data['max_lessons_per_day'])
        ]
    return schedule


async def fill_schedule(teachers, classes, schedule):
    """Алгоритм заполнения расписания"""
    all_placed = True

    for class_name, class_data in classes.items():
        total_lessons = sum(class_data['lessons'].values())
        max_per_day = class_data['max_lessons_per_day']

        if total_lessons > max_per_day * 5:
            logger.error(f"Невозможно распределить {total_lessons} уроков для {class_name}")
            all_placed = False
            continue

        for subject, count in class_data['lessons'].items():
            if count <= 0:
                continue

            # Поиск подходящего учителя
            teacher = await find_teacher_for_subject(subject, teachers)
            if not teacher:
                logger.warning(f"Не найден учитель для {subject}")
                all_placed = False
                continue

            # Распределение уроков
            for _ in range(count):
                placed = False
                for day in random.sample(range(5), 5):
                    for lesson_num in range(max_per_day):
                        if (not teacher['schedule'][day][lesson_num]
                                and schedule[class_name][lesson_num][day] == ""):
                            schedule[class_name][lesson_num][day] = (
                                f"{teacher['fio']} ({subject})"
                            )
                            teacher['schedule'][day][lesson_num] = True
                            teacher['lessons_count'] += 1
                            placed = True
                            break
                    if placed:
                        break
                if not placed:
                    all_placed = False

    return schedule, all_placed


async def find_teacher_for_subject(subject, teachers):
    """Поиск учителя для предмета"""
    candidates = []
    for t in teachers.values():
        if subject in t['subjects']:
            candidates.append(t)

    if candidates:
        return min(candidates, key=lambda x: x['lessons_count'])
    return None


async def print_schedule_table(schedule, classes, filename):
    """Генерация файла с расписанием"""
    with open(filename, "w", encoding="utf-8") as f:
        for class_name in schedule:
            max_lessons = classes[class_name]['max_lessons_per_day']
            f.write(f"Расписание для {class_name}:\n")
            f.write("+" + "-" * 58 + "+\n")
            f.write("| Урок | Пн       | Вт       | Ср       | Чт       | Пт       |\n")
            f.write("+" + "-" * 58 + "+\n")

            for lesson_num in range(max_lessons):
                row = [f"| {lesson_num + 1:^4}"]
                for day in range(5):
                    cell = schedule[class_name][lesson_num][day] or "Свободно"
                    row.append(f"{cell:^9}")
                f.write("|".join(row) + "|\n")
                f.write("+" + "-" * 58 + "+\n")


async def new_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Полный сброс данных и состояний с гарантированным выходом из диалога"""
    user_id = update.effective_user.id

    # 1. Принудительный выход из активного ConversationHandler
    current_handler = context.user_data.get('__conversation__')
    if current_handler:
        await context.application.dispatcher.process_update(update)
        del context.user_data['__conversation__']

    # 2. Полная очистка контекста
    context.user_data.clear()

    # 3. Удаление всех данных пользователя из БД
    async with aiosqlite.connect(DATABASE_NAME) as conn:
        await conn.execute('DELETE FROM subjects WHERE user_id = ?', (user_id,))
        await conn.execute('DELETE FROM teachers WHERE user_id = ?', (user_id,))
        await conn.execute('DELETE FROM classes WHERE user_id = ?', (user_id,))
        await conn.commit()

    # 4. Подтверждение с инлайн-клавиатурой для перезапуска
    await update.message.reply_text(
        "♻️ Все данные и состояния полностью сброшены!\n"
        "Нажмите /setup для нового расписания",
        reply_markup=ReplyKeyboardRemove()
    )

    # 5. Явное завершение всех диалогов
    return ConversationHandler.END

def setup_handlers(application: Application):
    """Настройка обработчиков с правильным приоритетом"""
    # 1. Регистрируем /new первым
    application.add_handler(CommandHandler("new", new_command))

    # 2. ConversationHandler с обработкой /new в fallbacks
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('setup', setup)],
        states={
            SUBJECTS: [MessageHandler(filters.TEXT & ~filters.COMMAND, input_subjects)],
            INPUT_SUBJECT: [MessageHandler(filters.TEXT & ~filters.COMMAND, input_subjects_next)],
            TEACHERS: [MessageHandler(filters.TEXT & ~filters.COMMAND, input_teachers)],
            TEACHER_SUBJECTS: [MessageHandler(filters.TEXT & ~filters.COMMAND, input_teachers_subjects)],
            TEACHER_TIME: [MessageHandler(filters.TEXT & ~filters.COMMAND, input_teachers_time)],
            CLASSES: [MessageHandler(filters.TEXT & ~filters.COMMAND, input_classes)],
            CLASS_MAX_LESSONS: [MessageHandler(filters.TEXT & ~filters.COMMAND, input_classes_max_lessons)],
            CLASS_MAX_LESSONS_VALUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, input_classes_max_lessons_value)],
            SCHEDULE: [MessageHandler(filters.TEXT & ~filters.COMMAND, input_classes_groups)],
        },
        fallbacks=[
            CommandHandler('new', new_command),  # Перехват /new внутри диалога
            CommandHandler('cancel', lambda u,c: ConversationHandler.END)
        ],
        allow_reentry=True  # Разрешаем перезапуск диалога
    )
    application.add_handler(conv_handler)
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("schedule", schedule_command))
    application.add_handler(CommandHandler("test", test_command))
    application.add_handler(CommandHandler("admin", admin_command))
    application.add_handler(CommandHandler("restart", restart_command))
    application.add_handler(CommandHandler("crash", crash_command))

async def main():
    await init_db()
    application = Application.builder().token(TOKEN).build()
    setup_handlers(application)
    await application.run_polling()

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.DEBUG  # Временное включение детального логгирования
)

if __name__ == '__main__':
    application = Application.builder().token(TOKEN).build()
    setup_handlers(application)


    async def main():
        await init_db()
        await application.initialize()
        await application.start()
        await application.updater.start_polling()

        # Бесконечное ожидание с периодическими проверками
        while True:
            await asyncio.sleep(3600)


    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\nБот остановлен")
    finally:
        application.stop()